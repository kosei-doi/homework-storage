#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
光と波動 実験2 - Excel上に直接グラフを作成するスクリプト
openpyxlを使ってExcelファイル内にグラフを直接作成します
"""

import pandas as pd
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.text import RichText
import os

def load_experiment_data(file_path):
    """Excelファイルから実験データを読み込む"""
    experiments = {}
    sheet_names = ['実験2-1_グラフ', '実験2-2_グラフ', '実験2-3_グラフ']
    
    for sheet_name in sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=4)
        data = []
        
        for idx, row in df.iterrows():
            if len(row) > 1:
                angle = row.iloc[1] if pd.notna(row.iloc[1]) else None
                
                if angle is not None:
                    try:
                        angle = float(angle)
                        if 0 <= angle <= 180:
                            measured = row.iloc[2] if len(row) > 2 else None
                            relative_intensity = row.iloc[4] if len(row) > 4 else None
                            theoretical = row.iloc[5] if len(row) > 5 else None
                            
                            measured = measured if pd.notna(measured) else None
                            relative_intensity = relative_intensity if pd.notna(relative_intensity) else None
                            theoretical = theoretical if pd.notna(theoretical) else None
                            
                            data.append({
                                'angle': angle,
                                'measured': measured,
                                'relative_intensity': relative_intensity,
                                'theoretical': theoretical
                            })
                    except (ValueError, TypeError):
                        continue
        
        experiments[sheet_name] = pd.DataFrame(data).dropna(subset=['angle'])
    
    return experiments

def create_chart_data_sheet(workbook, experiments):
    """グラフ用のデータシートを作成"""
    if 'グラフ用データ' in workbook.sheetnames:
        del workbook['グラフ用データ']
    
    ws_data = workbook.create_sheet('グラフ用データ', 0)
    
    # 全ての角度を集約
    all_angles = set()
    for df in experiments.values():
        all_angles.update(df['angle'].tolist())
    sorted_angles = sorted(all_angles)
    
    # ヘッダー行
    headers = ['角度 [degree]']
    col_map = {}  # シート名と列番号のマッピング
    
    col = 2
    experiment_labels = {
        '実験2-1_グラフ': ('実験2-1_測定値', '実験2-1_理論値'),
        '実験2-2_グラフ': ('実験2-2_測定値', '実験2-2_理論値'),
        '実験2-3_グラフ': ('実験2-3_測定値', '実験2-3_理論値'),
    }
    
    for sheet_name in ['実験2-1_グラフ', '実験2-2_グラフ', '実験2-3_グラフ']:
        if sheet_name in experiments:
            measured_label, theory_label = experiment_labels[sheet_name]
            headers.extend([measured_label, theory_label])
            col_map[sheet_name] = {'measured': col, 'theory': col + 1}
            col += 2
    
    ws_data.append(headers)
    
    # データ行
    for angle in sorted_angles:
        row = [angle]
        
        for sheet_name in ['実験2-1_グラフ', '実験2-2_グラフ', '実験2-3_グラフ']:
            if sheet_name in experiments:
                df = experiments[sheet_name]
                angle_data = df[df['angle'] == angle]
                
                if len(angle_data) > 0:
                    measured = angle_data.iloc[0]['relative_intensity']
                    theoretical = angle_data.iloc[0]['theoretical']
                    
                    row.append(measured if pd.notna(measured) else None)
                    row.append(theoretical if pd.notna(theoretical) and theoretical > 0 else None)
                else:
                    row.extend([None, None])
        
        ws_data.append(row)
    
    return ws_data, col_map

def create_chart_in_excel(workbook, ws_data, col_map):
    """Excel内にグラフを作成"""
    if 'グラフ' in workbook.sheetnames:
        del workbook['グラフ']
    
    ws_chart = workbook.create_sheet('グラフ')
    
    # 散布図チャートを作成
    chart = ScatterChart()
    chart.title = "光と波動 実験2 白色光の偏光特性"
    chart.style = 10
    chart.x_axis.title = '角度 θ [degree]'
    chart.y_axis.title = '相対強度'
    chart.height = 15
    chart.width = 24
    chart.legend.position = 'r'  # 右側に凡例
    
    max_row = ws_data.max_row
    
    # X軸データ（角度列）
    xvalues = Reference(ws_data, min_col=1, min_row=2, max_row=max_row)
    
    # 色とマーカーの設定
    style_config = {
        '実験2-1_グラフ': {'color': '0066CC', 'marker': 'circle', 'name': '実験2-1: 偏光板2枚'},
        '実験2-2_グラフ': {'color': 'CC0000', 'marker': 'square', 'name': '実験2-2: 偏光板3枚 (α=0°)'},
        '実験2-3_グラフ': {'color': '009900', 'marker': 'triangle', 'name': '実験2-3: 偏光板3枚 (α=10°)'},
    }
    
    # 各実験のデータ系列を追加
    for sheet_name in ['実験2-1_グラフ', '実験2-2_グラフ', '実験2-3_グラフ']:
        if sheet_name not in col_map:
            continue
        
        config = style_config[sheet_name]
        measured_col = col_map[sheet_name]['measured']
        theory_col = col_map[sheet_name]['theory']
        
        # 測定値の系列
        yvalues_measured = Reference(ws_data, min_col=measured_col, min_row=1, max_row=max_row)
        series_measured = Series(yvalues_measured, xvalues, title=config['name'] + ' (測定値)')
        series_measured.graphicalProperties.line.width = 30000  # 線の太さ
        series_measured.graphicalProperties.line.solidFill = config['color']
        chart.series.append(series_measured)
        
        # 理論値の系列
        yvalues_theory = Reference(ws_data, min_col=theory_col, min_row=1, max_row=max_row)
        series_theory = Series(yvalues_theory, xvalues, title=config['name'] + ' (理論値)')
        series_theory.graphicalProperties.line.width = 20000
        series_theory.graphicalProperties.line.solidFill = config['color']
        series_theory.graphicalProperties.line.dashStyle = 'sysDash'  # 破線
        chart.series.append(series_theory)
    
    # グリッド線を表示
    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines = ChartLines()
    
    # グラフをシートに追加
    ws_chart.add_chart(chart, "A1")
    
    return ws_chart

if __name__ == '__main__':
    file_path = '光と波動_実験2_グラフ.xlsx'
    output_path = '光と波動_実験2_グラフ_チャート付き.xlsx'
    
    print("=" * 60)
    print("光と波動 実験2 - Excelグラフ作成スクリプト")
    print("=" * 60)
    
    print("\n1. データを読み込んでいます...")
    experiments = load_experiment_data(file_path)
    
    for sheet_name, df in experiments.items():
        print(f"   {sheet_name}: {len(df)} データポイント")
    
    print("\n2. Excelファイルを開いています...")
    workbook = openpyxl.load_workbook(file_path)
    
    print("3. グラフ用データシートを作成しています...")
    ws_data, col_map = create_chart_data_sheet(workbook, experiments)
    
    print("4. グラフを作成しています...")
    ws_chart = create_chart_in_excel(workbook, ws_data, col_map)
    
    print(f"\n5. Excelファイルに保存しています...")
    workbook.save(output_path)
    
    print("\n" + "=" * 60)
    print("完了しました！")
    print("=" * 60)
    print(f"\n出力ファイル: {output_path}")
    print("\nファイルには以下のシートが含まれます:")
    print("  - グラフ用データ: グラフ作成用に整理されたデータ")
    print("  - グラフ: 自動生成されたグラフ")
    print("  - 元の実験シート（実験2-1_グラフ、実験2-2_グラフ、実験2-3_グラフ）")
    print("\nExcelで開いて、グラフの調整を行ってください。")

