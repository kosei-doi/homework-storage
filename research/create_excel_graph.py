#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
光と波動 実験2 - Excel上でグラフを作成するスクリプト
既存のExcelファイルにグラフ用データシートを追加し、グラフを作成します
"""

import pandas as pd
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.colors import ColorChoice
import os

def load_experiment_data(file_path):
    """Excelファイルから実験データを読み込む"""
    experiments = {}
    sheet_names = ['実験2-1_グラフ', '実験2-2_グラフ', '実験2-3_グラフ']
    
    for sheet_name in sheet_names:
        # ヘッダー行をスキップしてデータを読み込む
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=4)
        
        data = []
        for idx, row in df.iterrows():
            if len(row) > 1:
                angle = row.iloc[1] if pd.notna(row.iloc[1]) else None
                
                if angle is not None:
                    try:
                        angle = float(angle)
                        if 0 <= angle <= 180:
                            measured = row.iloc[2] if len(row) > 2 else None
                            relative_intensity = row.iloc[4] if len(row) > 4 else None
                            theoretical = row.iloc[5] if len(row) > 5 else None
                            
                            measured = measured if pd.notna(measured) else None
                            relative_intensity = relative_intensity if pd.notna(relative_intensity) else None
                            theoretical = theoretical if pd.notna(theoretical) else None
                            
                            data.append({
                                'angle': angle,
                                'measured': measured,
                                'relative_intensity': relative_intensity,
                                'theoretical': theoretical
                            })
                    except (ValueError, TypeError):
                        continue
        
        experiments[sheet_name] = pd.DataFrame(data).dropna(subset=['angle'])
    
    return experiments

def create_graph_data_sheet(workbook, experiments):
    """グラフ用のデータシートを作成"""
    # 既存のシートがあれば削除
    if 'グラフ用データ' in workbook.sheetnames:
        del workbook['グラフ用データ']
    
    ws_data = workbook.create_sheet('グラフ用データ', 0)
    
    # ヘッダー行を作成
    headers = ['角度 [degree]']
    
    experiment_configs = {
        '実験2-1_グラフ': '実験2-1_測定値',
        '実験2-2_グラフ': '実験2-2_測定値',
        '実験2-3_グラフ': '実験2-3_測定値',
    }
    
    theory_configs = {
        '実験2-1_グラフ': '実験2-1_理論値',
        '実験2-2_グラフ': '実験2-2_理論値',
        '実験2-3_グラフ': '実験2-3_理論値',
    }
    
    for sheet_name in ['実験2-1_グラフ', '実験2-2_グラフ', '実験2-3_グラフ']:
        if sheet_name in experiments:
            headers.append(experiment_configs[sheet_name])
            headers.append(theory_configs[sheet_name])
    
    ws_data.append(headers)
    
    # 全ての角度を集約
    all_angles = set()
    for df in experiments.values():
        all_angles.update(df['angle'].tolist())
    
    sorted_angles = sorted(all_angles)
    
    # データ行を作成
    for angle in sorted_angles:
        row = [angle]
        
        for sheet_name in ['実験2-1_グラフ', '実験2-2_グラフ', '実験2-3_グラフ']:
            if sheet_name in experiments:
                df = experiments[sheet_name]
                # 該当する角度のデータを取得
                angle_data = df[df['angle'] == angle]
                
                if len(angle_data) > 0:
                    measured = angle_data.iloc[0]['relative_intensity']
                    theoretical = angle_data.iloc[0]['theoretical']
                    
                    row.append(measured if pd.notna(measured) else '')
                    row.append(theoretical if pd.notna(theoretical) else '')
                else:
                    row.append('')
                    row.append('')
        
        ws_data.append(row)
    
    return ws_data

def create_chart_in_excel(workbook, ws_data):
    """Excel内にグラフを作成"""
    # グラフシートを作成（既存があれば削除）
    if 'グラフ' in workbook.sheetnames:
        del workbook['グラフ']
    
    ws_chart = workbook.create_sheet('グラフ')
    
    # 散布図チャートを作成
    chart = ScatterChart()
    chart.title = "光と波動 実験2 白色光の偏光特性"
    chart.style = 10
    chart.x_axis.title = '角度 θ [degree]'
    chart.y_axis.title = '相対強度'
    chart.height = 15
    chart.width = 20
    
    # データの範囲を定義
    max_row = ws_data.max_row
    max_col = ws_data.max_column
    
    # X軸データ（角度）
    xvalues = Reference(ws_data, min_col=1, min_row=2, max_row=max_row)
    
    # 各実験のデータ系列を追加
    colors = {
        '実験2-1_測定値': '0066CC',  # 青
        '実験2-1_理論値': '0066CC',
        '実験2-2_測定値': 'CC0000',  # 赤
        '実験2-2_理論値': 'CC0000',
        '実験2-3_測定値': '009900',  # 緑
        '実験2-3_理論値': '009900',
    }
    
    markers = {
        '実験2-1_測定値': 'circle',
        '実験2-2_測定値': 'square',
        '実験2-3_測定値': 'triangle',
    }
    
    # 列インデックスを取得
    for col_idx in range(2, max_col + 1, 2):  # 測定値列（2, 4, 6...）
        header_cell = ws_data.cell(row=1, column=col_idx)
        if header_cell.value:
            series_name = header_cell.value
            
            # 測定値の系列
            yvalues_measured = Reference(ws_data, min_col=col_idx, min_row=2, max_row=max_row)
            series = Series(yvalues_measured, xvalues, title=series_name)
            
            # マーカースタイルを設定
            if series_name in markers:
                for point in series.graphicalProperties.line:
                    pass  # マーカーの設定は後で行う
            
            chart.series.append(series)
            
            # 理論値の系列（次の列）
            if col_idx + 1 <= max_col:
                theory_header = ws_data.cell(row=1, column=col_idx + 1)
                if theory_header.value:
                    yvalues_theory = Reference(ws_data, min_col=col_idx + 1, min_row=2, max_row=max_row)
                    theory_series = Series(yvalues_theory, xvalues, title=theory_header.value)
                    chart.series.append(theory_series)
    
    # グラフをシートに追加
    ws_chart.add_chart(chart, "A1")
    
    return ws_chart

def create_simple_chart_data(workbook, experiments):
    """シンプルなグラフデータシートを作成（Excelの組み込み機能でグラフ作成可能）"""
    if 'グラフ用データ' in workbook.sheetnames:
        del workbook['グラフ用データ']
    
    ws_data = workbook.create_sheet('グラフ用データ', 0)
    
    # タイトル行
    ws_data['A1'] = '角度 [degree]'
    col = 2
    
    experiment_labels = {
        '実験2-1_グラフ': '実験2-1: 偏光板2枚 (測定値)',
        '実験2-2_グラフ': '実験2-2: 偏光板3枚 α=0° (測定値)',
        '実験2-3_グラフ': '実験2-3: 偏光板3枚 α=10° (測定値)',
    }
    
    theory_labels = {
        '実験2-1_グラフ': '実験2-1 (理論値)',
        '実験2-2_グラフ': '実験2-2 (理論値)',
        '実験2-3_グラフ': '実験2-3 (理論値)',
    }
    
    # 全ての角度を集約
    all_angles = set()
    for df in experiments.values():
        all_angles.update(df['angle'].tolist())
    sorted_angles = sorted(all_angles)
    
    # 各実験のデータを列として配置
    for sheet_name in ['実験2-1_グラフ', '実験2-2_グラフ', '実験2-3_グラフ']:
        if sheet_name in experiments:
            df = experiments[sheet_name]
            
            # 測定値列
            ws_data.cell(row=1, column=col).value = experiment_labels[sheet_name]
            for row_idx, angle in enumerate(sorted_angles, start=2):
                angle_data = df[df['angle'] == angle]
                if len(angle_data) > 0:
                    value = angle_data.iloc[0]['relative_intensity']
                    if pd.notna(value):
                        ws_data.cell(row=row_idx, column=1).value = angle
                        ws_data.cell(row=row_idx, column=col).value = value
            
            col += 1
            
            # 理論値列
            ws_data.cell(row=1, column=col).value = theory_labels[sheet_name]
            for row_idx, angle in enumerate(sorted_angles, start=2):
                angle_data = df[df['angle'] == angle]
                if len(angle_data) > 0:
                    value = angle_data.iloc[0]['theoretical']
                    if pd.notna(value) and value > 0:
                        ws_data.cell(row=row_idx, column=col).value = value
            
            col += 1
    
    return ws_data

if __name__ == '__main__':
    file_path = '光と波動_実験2_グラフ.xlsx'
    output_path = '光と波動_実験2_グラフ_チャート付き.xlsx'
    
    print("データを読み込んでいます...")
    experiments = load_experiment_data(file_path)
    
    # 各実験のデータを確認
    for sheet_name, df in experiments.items():
        print(f"\n{sheet_name}:")
        print(f"  データ点数: {len(df)}")
        print(f"  角度範囲: {df['angle'].min():.1f}° ～ {df['angle'].max():.1f}°")
    
    print("\nExcelファイルを開いています...")
    # 既存のファイルを読み込む
    workbook = openpyxl.load_workbook(file_path)
    
    print("グラフ用データシートを作成しています...")
    ws_data = create_simple_chart_data(workbook, experiments)
    
    print("Excelファイルに保存しています...")
    workbook.save(output_path)
    
    print(f"\n完了しました！")
    print(f"出力ファイル: {output_path}")
    print(f"\n次のステップ:")
    print(f"1. '{output_path}' を開く")
    print(f"2. 'グラフ用データ' シートを選択")
    print(f"3. データ範囲を選択（A1から全てのデータ列まで）")
    print(f"4. 挿入 → グラフ → 散布図（マーカーと線）を選択")
    print(f"5. グラフのタイトルや軸ラベルを設定")
    print(f"\nまたは、Pythonで自動的にグラフを作成することもできます。")

